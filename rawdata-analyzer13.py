import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------- Load Excel ----------------------
file_path = r'C:/Users/Laptop-SatyoYuwono/Downloads/RD-ZS.xlsx'
df = pd.read_excel(file_path, header=None)

# ---------------------- Column Index Map ----------------------
col_map = {
    'vessel_code': 0,
    'telegram_date': 1,
    'telegram_type': 2,
    'me_hsfo': 22,
    'me_lsfo': 23,
    'ae_hsfo': 26,
    'ae_lsfo': 27,
    'boiler_hsfo': 30,
    'boiler_lsfo': 31,
    'fw_rob': 57,
    'fw_prod': 58,
    'fw_bunk': 59,
    'engine_rpm': 66,
    'prop_pitch': 65,
    'miles_slc': 116,
    'hours_slc': 65,
    'minutes_slc': 66,
    'wind_course': 84,
    'wind_force': 85,
    'swell_course': 90,
    'swell_force': 91,
    'current_course': 88,
    'current_force': 89,
    'vessel_course': 92,
    'come_cons': 49,
    'supplied_co': 54,

    'ae1_kw': 79,
    'ae2_kw': 81,
    'ae3_kw': 83
}

# Extract relevant columns
df_data = df.iloc[:, list(col_map.values())].copy()
df_data.columns = list(col_map.keys())
df_data['telegram_date'] = pd.to_datetime(df_data['telegram_date'], errors='coerce')

# ---------------------- Date Filter ----------------------
voyage_start_date = pd.to_datetime('2025-01-07')
voyage_end_date = pd.to_datetime('2025-01-26')
df_data = df_data[(df_data['telegram_date'] >= voyage_start_date) & (df_data['telegram_date'] <= voyage_end_date)]

# ---------------------- Calculations ----------------------
df_data['daily_fo_cons'] = df_data[['me_hsfo', 'me_lsfo', 'ae_hsfo', 'ae_lsfo', 'boiler_hsfo', 'boiler_lsfo']].sum(axis=1)
df_data['fw_consumption'] = df_data['fw_rob'].shift(1) + df_data['fw_bunk'] + df_data['fw_prod'] - df_data['fw_rob']
df_data['fw_production'] = df_data['fw_prod']
df_data['total_hrs'] = df_data['hours_slc'] + (df_data['minutes_slc'] / 60)

# Only keep rows with positive total_hrs to avoid division by zero
df_data = df_data[df_data['total_hrs'] > 0].copy()

df_data['actual_speed'] = df_data.apply(
    lambda row: row['miles_slc'] / row['total_hrs'] if pd.notna(row['total_hrs']) and row['total_hrs'] > 0 else np.nan,
    axis=1
)
df_data['slip'] = df_data.apply(
    lambda row: 1 - ((row['actual_speed'] * 1852) / (row['engine_rpm'] * row['prop_pitch'] * 60))
    if pd.notna(row['engine_rpm']) and pd.notna(row['prop_pitch']) and row['engine_rpm'] > 0 and row['prop_pitch'] > 0 else np.nan,
    axis=1
)
df_data['performance_speed'] = (df_data['engine_rpm'] * df_data['prop_pitch'] * 60) / 1852 # ideal condition
df_data['come_cons_delta'] = df_data['come_cons'].shift(1) + df_data['supplied_co'] - df_data['come_cons']

# ---------------------- Convert Weather Columns ----------------------
for col in ['wind_force', 'vessel_course', 'wind_course', 'swell_force', 'swell_course', 'current_force', 'current_course']:
    df_data[col] = pd.to_numeric(df_data[col], errors='coerce')

# ---------------------- Weather Speed Loss Mapping ----------------------
def calc_relative_effect(vessel_course, weather_course):
    try:
        angle = abs(float(vessel_course) - float(weather_course)) % 360
        if angle > 180:
            angle = 360 - angle
        return 1 if angle < 90 else -1
    except:
        return 0

def wind_speed_loss(force):
    loss_map = {0: 0.00, 1: 0.00, 2: 0.00, 3: 0.015, 4: 0.03, 5: 0.05, 6: 0.075, 7: 0.10, 8: 0.13, 9: 0.17, 10: 0.22, 11: 0.28, 12: 0.35}
    return loss_map.get(int(force), 0.0)

def swell_speed_loss(force):
    if force <= 0.5: return 0.00
    elif force <= 1.0: return 0.01
    elif force <= 2.0: return 0.025
    elif force <= 3.0: return 0.05
    elif force <= 4.0: return 0.08
    elif force <= 5.0: return 0.12
    else: return 0.15

def current_speed_loss(force):
    if force <= 0.2: return 0.00
    elif force <= 0.5: return 0.005
    elif force <= 1.0: return 0.01
    elif force <= 1.5: return 0.02
    elif force <= 2.0: return 0.035
    else: return 0.05

def safe_calc_loss(row, force_col, course_col, loss_func):
    try:
        force = float(row[force_col])
        ps = float(row['performance_speed'])
        vc = float(row['vessel_course'])
        wc = float(row[course_col])
        direction = calc_relative_effect(vc, wc)
        return loss_func(force) * ps * direction
    except:
        return np.nan

df_data['wind_loss'] = df_data.apply(lambda r: safe_calc_loss(r, 'wind_force', 'wind_course', wind_speed_loss), axis=1)
df_data['swell_loss'] = df_data.apply(lambda r: safe_calc_loss(r, 'swell_force', 'swell_course', swell_speed_loss), axis=1)
df_data['current_loss'] = df_data.apply(lambda r: safe_calc_loss(r, 'current_force', 'current_course', current_speed_loss), axis=1)

# ---------------------- Save to Excel ----------------------
output_path = 'C:/Users/Laptop-SatyoYuwono/Downloads/NY-VOYAGE-REV16.xlsx'
wb = openpyxl.Workbook()

# ---- Sheet 1: General Data ----
ws1 = wb.active
ws1.title = "General Data"

table1 = df_data[['telegram_date', 'daily_fo_cons', 'fw_consumption', 'fw_production', 'performance_speed', 'actual_speed']]
for r in dataframe_to_rows(table1, index=False, header=True):
    ws1.append(r)

ws1.append([])
ws1.append(["Date", "Speed Lost by Current", "Speed Lost by Swell", "Speed Lost by Wind"])
for i in range(len(df_data)):
    ws1.append([df_data['telegram_date'].iloc[i], df_data['current_loss'].iloc[i], df_data['swell_loss'].iloc[i], df_data['wind_loss'].iloc[i]])

bar1 = BarChart()
bar1.title = "Fuel, FW, and Speed Performance"
bar1.add_data(Reference(ws1, min_col=2, min_row=2, max_col=6, max_row=1+len(table1)), titles_from_data=True)
bar1.set_categories(Reference(ws1, min_col=1, min_row=2, max_row=1+len(table1)))
ws1.add_chart(bar1, "I2")

start_row = len(table1) + 5
bar2 = BarChart()
bar2.title = "Speed Loss by Weather"
bar2.add_data(Reference(ws1, min_col=1, min_row=start_row, max_col=3, max_row=start_row+len(df_data)-1), titles_from_data=False)
ws1.add_chart(bar2, "I20")

# ---- Sheet 2: Main Engine Data ----
ws2 = wb.create_sheet("Main Engine Data")
me_data = df_data[['telegram_date', 'engine_rpm', 'slip', 'me_hsfo', 'me_lsfo', 'come_cons_delta']].copy()
me_data['me_fo_total'] = me_data['me_hsfo'] + me_data['me_lsfo']
me_table = me_data[['telegram_date', 'engine_rpm', 'slip', 'me_fo_total', 'come_cons_delta']]
for r in dataframe_to_rows(me_table, index=False, header=True):
    ws2.append(r)

bar3 = BarChart()
bar3.title = "Main Engine Metrics"
bar3.add_data(Reference(ws2, min_col=2, min_row=2, max_col=5, max_row=1+len(me_table)), titles_from_data=True)
bar3.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=1+len(me_table)))
ws2.add_chart(bar3, "H2")

# ---- Sheet 3: Aux. Engine & Boiler ----
ws3 = wb.create_sheet("Aux. Engine & Boiler Data")
aux_data = df_data[['telegram_date', 'boiler_hsfo', 'boiler_lsfo', 'ae_hsfo', 'ae_lsfo', 'ae1_kw', 'ae2_kw', 'ae3_kw']].copy()
aux_data['boiler_fo'] = aux_data['boiler_hsfo'] + aux_data['boiler_lsfo']
aux_data['ae_fo'] = aux_data['ae_hsfo'] + aux_data['ae_lsfo']
aux_table = aux_data[['telegram_date', 'boiler_fo', 'ae_fo', 'ae1_kw', 'ae2_kw', 'ae3_kw']]
for r in dataframe_to_rows(aux_table, index=False, header=True):
    ws3.append(r)

bar4 = BarChart()
bar4.title = "Aux Engine and Boiler Metrics"
bar4.add_data(Reference(ws3, min_col=2, min_row=2, max_col=6, max_row=1+len(aux_table)), titles_from_data=True)
bar4.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=1+len(aux_table)))
ws3.add_chart(bar4, "I2")

# Save
wb.save(output_path)
print("Done: NY Voyage Data Rev.13 processed and saved.")