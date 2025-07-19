import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------- Load Excel ----------------------
file_path = r'C:/Users/Laptop-SatyoYuwono/Downloads/RD-ZS.xlsx'
df = pd.read_excel(file_path, header=None)

# ---------------------- Column Index Map ----------------------
col_map = {
    'vessel_code': 0,
    'telegram_date': 1,
    'telegram_type': 2,
    'me_hsfo': 22,
    'me_lsfo': 23,
    'ae_hsfo': 26,
    'ae_lsfo': 27,
    'boiler_hsfo': 30,
    'boiler_lsfo': 31,
    'fw_rob': 57,
    'fw_prod': 58,
    'fw_bunk': 59,
    'engine_rpm': 66,
    'prop_pitch': 65,
    'miles_slc': 62,
    'hours_slc': 63,
    'minutes_slc': 64,
    'wind_course': 84,
    'wind_force': 85,
    'swell_course': 90,
    'swell_force': 91,
    'current_course': 88,
    'current_force': 89,
    'vessel_course': 92,
    'come_cons': 49,
    'supplied_co': 54,
    'ae1_hours': 80,
    'ae1_kw': 79,
    'ae2_hours': 82,
    'ae2_kw': 81,
    'ae3_hours': 84,
    'ae3_kw': 83
}

# Extract and rename relevant columns
df_data = df.iloc[:, list(col_map.values())].copy()
df_data.columns = list(col_map.keys())

# --- Specify date format for telegram_date ---
df_data['telegram_date'] = pd.to_datetime(df_data['telegram_date'], errors='coerce')

# **IMPORTANT: Sort df_data by date FIRST to ensure correct shifting**
df_data = df_data.sort_values(by='telegram_date').reset_index(drop=True)

# --- Convert relevant columns to numeric and handle NaNs ---
numeric_cols = [
    'me_hsfo', 'me_lsfo', 'ae_hsfo', 'ae_lsfo', 'boiler_hsfo', 'boiler_lsfo',
    'fw_rob', 'fw_prod', 'fw_bunk',
    'engine_rpm', 'prop_pitch', 'miles_slc', 'hours_slc', 'minutes_slc',
    'come_cons', 'supplied_co', 'ae1_kw', 'ae2_kw', 'ae3_kw',
    'wind_force', 'vessel_course', 'wind_course', 'swell_force', 'swell_course', 'current_force', 'current_course'
]

for col in numeric_cols:
    df_data[col] = pd.to_numeric(df_data[col], errors='coerce').fillna(0)

# ---------------------- Calculations ----------------------
df_data['Daily FO Consumption (kL)'] = df_data[['me_hsfo', 'me_lsfo', 'ae_hsfo', 'ae_lsfo', 'boiler_hsfo', 'boiler_lsfo']].sum(axis=1)
df_data['Daily FW Consumption (kL)'] = df_data['fw_rob'].shift(1) + df_data['fw_bunk'] + df_data['fw_prod'] - df_data['fw_rob']
df_data['Daily FW Production (kL)'] = df_data['fw_prod']
df_data['Engine RPM'] = df_data['engine_rpm']
# This line already calculates ME FO Consumption (kL)
df_data['ME FO Consumption (kL)'] = df_data['me_hsfo'] + df_data['me_lsfo']
df_data['total_hrs'] = df_data['hours_slc'] + (df_data['minutes_slc'] / 60)
df_data['AE 1 Power (kW)'] = df_data['ae1_kw']
df_data['AE 2 Power (kW)'] = df_data['ae2_kw']
df_data['AE 3 Power (kW)'] = df_data['ae3_kw']
df_data['AE 1 Hours'] = df_data['ae1_hours']
df_data['AE 2 Hours'] = df_data['ae2_hours']
df_data['AE 3 Hours'] = df_data['ae3_hours']
df_data['COME Consumption (L)'] = (
    df_data['come_cons'].shift(1).fillna(0) - df_data['come_cons'].fillna(0)
) + df_data['supplied_co'].fillna(0)

# Also pre-calculate Boiler FO and AE FO Consumption here with descriptive names
df_data['Boiler FO Consumption (kL)'] = df_data['boiler_hsfo'] + df_data['boiler_lsfo']
df_data['AE FO Consumption (kL)'] = df_data['ae_hsfo'] + df_data['ae_lsfo']


# ---------------------- Date Filter ----------------------
voyage_start_date = pd.to_datetime('2025-01-14')
voyage_end_date = pd.to_datetime('2025-03-06')
df_data = df_data[(df_data['telegram_date'] >= voyage_start_date) & (df_data['telegram_date'] <= voyage_end_date)].copy()

# Filter rows with positive total_hrs (after date filter)
df_data = df_data[df_data['total_hrs'] > 0].copy()

# Calculate speeds and slip (these don't rely on previous rows, so fine here)
df_data['Actual Speed (knots)'] = df_data['miles_slc'] / df_data['total_hrs']
df_data['Slip (%)'] = df_data.apply(
    lambda row: ((1 - ((row['Actual Speed (knots)'] * 1852) / (row['engine_rpm'] * row['prop_pitch'] * 60))) * 100)
    if row['engine_rpm'] > 0 and row['prop_pitch'] > 0 else np.nan,
    axis=1
)
df_data['Performance Speed (knots)'] = (df_data['engine_rpm'] * df_data['prop_pitch'] * 60) / 1852

# ---------------------- Weather Speed Loss ----------------------
# Functions defined as before

def calc_relative_effect(vessel_course, weather_course):
    try:
        angle = abs(float(vessel_course) - float(weather_course)) % 360
        if angle > 180:
            angle = 360 - angle
        return 1 if angle < 90 else -1
    except:
        return 0

def wind_speed_loss(force):
    loss_map = {0: 0.00, 1: 0.00, 2: 0.00, 3: 0.015, 4: 0.03, 5: 0.05, 6: 0.075, 7: 0.10, 8: 0.13, 9: 0.17, 10: 0.22, 11: 0.28, 12: 0.35}
    return loss_map.get(int(force), 0.0)

def swell_speed_loss(force):
    if force <= 0.5: return 0.00
    elif force <= 1.0: return 0.01
    elif force <= 2.0: return 0.025
    elif force <= 3.0: return 0.05
    elif force <= 4.0: return 0.08
    elif force <= 5.0: return 0.12
    else: return 0.15

def current_speed_loss(force):
    if force <= 0.2: return 0.00
    elif force <= 0.5: return 0.005
    elif force <= 1.0: return 0.01
    elif force <= 1.5: return 0.02
    elif force <= 2.0: return 0.035
    else: return 0.05

def safe_calc_loss(row, force_col, course_col, loss_func):
    try:
        force = float(row[force_col])
        ps = float(row['Performance Speed (knots)'])
        vc = float(row['vessel_course'])
        wc = float(row[course_col])
        direction = calc_relative_effect(vc, wc)
        return loss_func(force) * ps * direction
    except:
        return np.nan

df_data['Speed Loss by Wind (knots)'] = df_data.apply(lambda r: safe_calc_loss(r, 'wind_force', 'wind_course', wind_speed_loss), axis=1)
df_data['Speed Loss by Swell (knots)'] = df_data.apply(lambda r: safe_calc_loss(r, 'swell_force', 'swell_course', swell_speed_loss), axis=1)
df_data['Speed Loss by Current (knots)'] = df_data.apply(lambda r: safe_calc_loss(r, 'current_force', 'current_course', current_speed_loss), axis=1)

# ---------------------- Prepare for Excel ----------------------
df_filtered_hrs = df_data[df_data['total_hrs'] > 10].copy()
df_filtered_hrs = df_filtered_hrs.reset_index(drop=True)

output_path = 'C:/Users/Laptop-SatyoYuwono/Downloads/EVR-AA-VOY1.xlsx'
wb = openpyxl.Workbook()

# ---- Sheet 1 ----
ws1 = wb.active
ws1.title = "General Data"
table1 = df_filtered_hrs[['telegram_date', 'Daily FO Consumption (kL)', 'Daily FW Consumption (kL)', 'Daily FW Production (kL)', 'Performance Speed (knots)', 'Actual Speed (knots)']]
for r in dataframe_to_rows(table1, index=False, header=True):
    ws1.append(r)

bar1 = BarChart()
bar1.title = "Fuel, FW, and Speed Performance"
# Corrected min_row to 2 for data/categories as headers are in row 1
bar1.add_data(Reference(ws1, min_col=2, min_row=1, max_col=6, max_row=1+len(table1)), titles_from_data=True)
bar1.set_categories(Reference(ws1, min_col=1, min_row=1, max_row=1+len(table1)))
ws1.add_chart(bar1, "I2")

# --- REVISED bar2 Script (Dates on X-axis) ---
ws1.append([])
header_for_bar2_data = ["Date", "Speed Loss by Current (knots)", "Speed Loss by Swell (knots)", "Speed Loss by Wind (knots)"]
ws1.append(header_for_bar2_data)

data_start_row_for_bar2 = ws1.max_row + 1

for i in range(len(df_filtered_hrs)):
    ws1.append([
        df_filtered_hrs['telegram_date'].iloc[i],
        df_filtered_hrs['Speed Loss by Current (knots)'].iloc[i],
        df_filtered_hrs['Speed Loss by Swell (knots)'].iloc[i],
        df_filtered_hrs['Speed Loss by Wind (knots)'].iloc[i]
    ])

bar2 = BarChart()
bar2.title = "Speed Loss by Weather"

# Define Categories (X-axis labels) - Use the Date column (Col 1)
# min_row points to the first data row, skipping the header for bar2
bar2.set_categories(Reference(ws1, min_col=1, min_row=data_start_row_for_bar2-1, max_row=data_start_row_for_bar2 + len(df_filtered_hrs) - 1))

# Define Data Series (Y-axis values) - Use speed loss columns (Col 2, 3, 4)
# titles_from_data=True will pick up "Speed Loss by Current (knots)", etc. from the header_for_bar2_data row.
bar2.add_data(Reference(ws1, min_col=2, min_row=data_start_row_for_bar2-1, max_col=4, max_row=data_start_row_for_bar2 + len(df_filtered_hrs) - 1), titles_from_data=True)

bar2.x_axis.title = "Date"
bar2.y_axis.title = "Speed Loss (Knots)"

ws1.add_chart(bar2, "I20")

# ---- Sheet 2 ----
ws2 = wb.create_sheet("Main Engine Data")
# Use the pre-calculated and renamed columns directly
me_data = df_filtered_hrs[['telegram_date', 'Engine RPM', 'Slip (%)', 'ME FO Consumption (kL)', 'COME Consumption (L)']].copy()
# Removed the redundant me_fo_total calculation on me_data
me_table = me_data[['telegram_date', 'Engine RPM', 'Slip (%)', 'ME FO Consumption (kL)', 'COME Consumption (L)']]
for r in dataframe_to_rows(me_table, index=False, header=True):
    ws2.append(r)

bar3 = BarChart()
bar3.title = "Main Engine Metrics"
# Corrected min_row to 2 for data/categories
bar3.add_data(Reference(ws2, min_col=2, min_row=1, max_col=5, max_row=1+len(me_table)), titles_from_data=True)
bar3.set_categories(Reference(ws2, min_col=1, min_row=1, max_row=1+len(me_table)))
ws2.add_chart(bar3, "H2")

# ---- Sheet 3 ----
ws3 = wb.create_sheet("Aux. Engine & Boiler Data")
# aux_data columns defined correctly and using pre-calculated FO consumption
aux_table = df_filtered_hrs[['telegram_date', 'Boiler FO Consumption (kL)', 'AE FO Consumption (kL)', 'AE 1 Power (kW)', 'AE 2 Power (kW)', 'AE 3 Power (kW)']]
for r in dataframe_to_rows(aux_table, index=False, header=True):
    ws3.append(r)

bar4 = BarChart()
bar4.title = "Aux Engine and Boiler Metrics"
# Corrected min_row to 2 for data/categories
bar4.add_data(Reference(ws3, min_col=2, min_row=1, max_col=6, max_row=1+len(aux_table)), titles_from_data=True)
bar4.set_categories(Reference(ws3, min_col=1, min_row=1, max_row=1+len(aux_table)))
ws3.add_chart(bar4, "I2")

# Save
wb.save(output_path)
print("Done: NY Voyage Data Rev.15 FINAL processed and saved.")